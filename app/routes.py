import holidays
import pytz
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from . import db
from .models import WorkHours, StudySchedule, CourseType
import datetime
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
import pandas as pd
import calendar
from openpyxl.styles import PatternFill
from .summary import get_summary_data
from reportlab.lib import colors

main = Blueprint('main', __name__)


@main.route('/')
@login_required
def index():
    today_date = datetime.date.today()
    work_record = WorkHours.query.filter_by(user_id=current_user.id, date=today_date, end_time=None).first()
    is_working = bool(work_record)

    warsaw_tz = pytz.timezone('Europe/Warsaw')
    current_time = datetime.datetime.now(warsaw_tz).strftime('%H:%M')

    course_types = CourseType.query.all()
    return render_template('index.html', is_working=is_working, current_time=current_time, course_types=course_types)


@main.route('/summary')
@login_required
def summary():
    month = request.args.get('month', type=int, default=datetime.date.today().month)
    year = request.args.get('year', type=int, default=datetime.date.today().year)

    summary_data = get_summary_data(month, year)
    return render_template(
        'summary.html',
        **summary_data
    )



@main.route('/begin_work', methods=['POST'])
@login_required
def begin_work():
    arrival_time = request.form.get('arrival_time')
    today_date = datetime.date.today()
    try:
        arrival_time_obj = datetime.datetime.strptime(arrival_time, '%H:%M').time()
    except ValueError:
        flash('Błędny format godziny', 'danger')
        return redirect(url_for('main.index'))

    existing_record = WorkHours.query.filter_by(user_id=current_user.id, date=today_date, end_time=None).first()
    if existing_record:
        flash('Praca już rozpoczęta.', 'danger')
        return redirect(url_for('main.index'))

    try:
        work_hours = WorkHours(user_id=current_user.id, date=today_date, start_time=arrival_time_obj)
        db.session.add(work_hours)
        db.session.commit()
        flash('Praca rozpoczęta pomyślnie.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Błąd: {str(e)}', 'danger')

    return redirect(url_for('main.index'))


@main.route('/end_work', methods=['POST'])
@login_required
def end_work():
    today_date = datetime.date.today()
    end_time = request.form.get('end_time')

    try:
        end_time_obj = datetime.datetime.strptime(end_time, '%H:%M').time()
    except ValueError:
        flash("Błędny format godziny", 'danger')
        return redirect(url_for('main.index'))

    work_record = WorkHours.query.filter_by(user_id=current_user.id, date=today_date, end_time=None).first()
    if not work_record:
        flash("Nie rozpoczęto pracy.", 'danger')
        return redirect(url_for('main.index'))

    try:
        work_record.end_time = end_time_obj
        db.session.commit()
        flash('Praca zakończona pomyślnie.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Błąd: {str(e)}", 'danger')

    return redirect(url_for('main.index'))


@main.route('/add_work_hours', methods=['POST'])
@login_required
def add_work_hours():
    date = request.form.get('date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')

    # Konwersja daty i godzin na obiekty czasowe
    try:
        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        start_time_obj = datetime.datetime.strptime(start_time, '%H:%M').time()
        end_time_obj = datetime.datetime.strptime(end_time, '%H:%M').time()
    except ValueError:
        flash("Nieprawidłowy format daty lub godziny.", 'danger')
        return redirect(url_for('main.index'))

    # Sprawdzenie, czy istnieje już nakładający się zakres godzin dla danego dnia
    overlapping_record = WorkHours.query.filter(
        WorkHours.user_id == current_user.id,
        WorkHours.date == date_obj,
        ((WorkHours.start_time < end_time_obj) & (WorkHours.end_time > start_time_obj)) |
        ((WorkHours.start_time < end_time_obj) & (WorkHours.end_time.is_(None)))
    ).first()

    if overlapping_record:
        flash("Podane godziny nakładają się na istniejący zakres godzin pracy.", 'danger')
        return redirect(url_for('main.index'))

    # Dodanie nowego rekordu godzin pracy, jeśli nie ma konfliktu
    try:
        work_hours = WorkHours(
            user_id=current_user.id,
            date=date_obj,
            start_time=start_time_obj,
            end_time=end_time_obj
        )
        db.session.add(work_hours)
        db.session.commit()
        flash('Godziny pracy zostały dodane.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Wystąpił błąd podczas dodawania godzin pracy: {str(e)}", 'danger')

    return redirect(url_for('main.index'))


@main.route('/add_study_schedule', methods=['POST'])
@login_required
def add_study_schedule():
    # Odczytanie wartości formularza
    date_range = request.form.get('date_range[]')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    subject_name = request.form.get('subject_name')
    instructor_name = request.form.get('instructor_name') or None
    room = request.form.get('room') or None
    week_type_id = request.form.get('week_type_id') or None
    course_type_id = request.form.get('course_type_id')
    day_of_week = request.form.get('day_of_week')

    # Pobranie pojedynczych dat z formularza
    single_dates = request.form.getlist('dates[]')

    errors = []

    # Walidacja danych wejściowych
    if not start_time or not end_time:
        errors.append("Proszę podać godziny rozpoczęcia i zakończenia zajęć.")
    if not subject_name:
        errors.append("Proszę podać nazwę przedmiotu.")
    if not course_type_id:
        errors.append("Proszę wybrać typ zajęć.")
    if date_range and not day_of_week:
        errors.append("Jeśli wybierzesz zakres dat, musisz również podać dzień tygodnia.")
    elif not date_range and not day_of_week and not single_dates:
        errors.append("Proszę podać przynajmniej jedną datę, dzień tygodnia lub pojedyncze daty.")

    if errors:
        return jsonify({'errors': errors}), 400

    try:
        records_added = 0

        # Walidacja zakresu dat
        if date_range:
            try:
                start_date_str, end_date_str = date_range.split(' - ')
                start_date = datetime.datetime.strptime(start_date_str.strip(), '%Y-%m-%d')
                end_date = datetime.datetime.strptime(end_date_str.strip(), '%Y-%m-%d')

                today = datetime.datetime.now()
                one_year_later = today + datetime.timedelta(days=365)
                one_year_earlier = today - datetime.timedelta(days=365)

                # Sprawdzenie, czy daty są w dozwolonym zakresie
                if start_date < one_year_earlier or end_date > one_year_later:
                    errors.append("Zakres dat nie może przekraczać jednego roku w przód ani w tył.")
                    return jsonify({'errors': errors}), 400
            except ValueError:
                errors.append("Nieprawidłowy format daty. Użyj formatu YYYY-MM-DD.")
                return jsonify({'errors': errors}), 400

        # Dodanie dat z zakresu
        if date_range and day_of_week:
            day_of_week_map = {
                "Poniedziałek": 0,
                "Wtorek": 1,
                "Środa": 2,
                "Czwartek": 3,
                "Piątek": 4,
                "Sobota": 5,
                "Niedziela": 6
            }
            target_weekday = day_of_week_map.get(day_of_week)

            current_date = start_date
            while current_date <= end_date:
                if current_date.weekday() == target_weekday:
                    if week_type_id:
                        week_number = current_date.isocalendar()[1]
                        is_even_week = (week_number % 2 == 0)
                        if (week_type_id == '1' and not is_even_week) or (week_type_id == '2' and is_even_week):
                            current_date += datetime.timedelta(days=1)
                            continue

                    study_schedule = StudySchedule(
                        user_id=current_user.id,
                        date=current_date.date(),
                        day_of_week=day_of_week,
                        start_time=start_time,
                        end_time=end_time,
                        subject_name=subject_name,
                        instructor_name=instructor_name,
                        room=room,
                        week_type_id=week_type_id,
                        course_type_id=course_type_id
                    )
                    db.session.add(study_schedule)
                    records_added += 1

                current_date += datetime.timedelta(days=1)

        # Dodanie pojedynczych dat
        for date_str in single_dates:
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
            study_schedule = StudySchedule(
                user_id=current_user.id,
                date=date.date(),
                day_of_week=date.strftime("%A"),
                start_time=start_time,
                end_time=end_time,
                subject_name=subject_name,
                instructor_name=instructor_name,
                room=room,
                week_type_id=week_type_id,
                course_type_id=course_type_id
            )
            db.session.add(study_schedule)
            records_added += 1

        db.session.commit()

        if records_added == 0:
            return jsonify(
                {'error': 'Nie dodano żadnych rekordów do harmonogramu. Upewnij się, że dane są poprawne.'}), 400

        return jsonify({'message': 'Harmonogram zajęć został dodany dla wybranych dat.'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f"Wystąpił błąd podczas dodawania harmonogramu: {str(e)}"}), 500


@main.route('/get_calendar_events')
@login_required
def get_calendar_events():
    try:
        # Parsowanie zakresu dat
        start = request.args.get('start')
        end = request.args.get('end')

        if start is None or end is None:
            return jsonify({"error": "Zakres dat nie został określony"}), 400

        # Przekonwertowanie na obiekt datetime.date
        start_date = datetime.datetime.strptime(start, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end, '%Y-%m-%d').date()

        # Pobierz wydarzenia w zakresie dat
        study_schedules = StudySchedule.query.filter(
            StudySchedule.user_id == current_user.id,
            StudySchedule.date >= start_date,
            StudySchedule.date <= end_date
        ).all()

        work_hours = WorkHours.query.filter(
            WorkHours.user_id == current_user.id,
            WorkHours.date >= start_date,
            WorkHours.date <= end_date,
            WorkHours.end_time.isnot(None)
        ).all()

        events = []
        course_colors = {
            "Ćwiczenia": "#FF5733",
            "Egzamin": "#C70039",
            "Kolokwium": "#900C3F",
            "Laboratorium": "#0733ad",
            "Projekt": "#d6207b",
            "Wykład": "#0fa835"
        }

        # Dodawanie wydarzeń związanych z harmonogramem zajęć
        for study in study_schedules:
            course_type = CourseType.query.get(study.course_type_id).name
            color = course_colors.get(course_type, "#808080")

            events.append({
                "id": str(study.id),
                "title": f"{study.subject_name} ({course_type})",
                "start": f"{study.date}T{study.start_time.isoformat()}",
                "end": f"{study.date}T{study.end_time.isoformat()}",
                "color": color,
                "textColor": "#FFFFFF",
                "extendedProps": {
                    "type": "study",
                    "date": study.date.isoformat(),
                    "start_time": study.start_time.isoformat(),
                    "end_time": study.end_time.isoformat(),
                    "subject_name": study.subject_name,
                    "course_type_id": study.course_type_id,
                    "course_type": course_type,
                    "day_of_week": getattr(study, "day_of_week", None),
                    "instructor_name": getattr(study, "instructor_name", None),
                    "room": getattr(study, "room", None),
                    "week_type_id": getattr(study, "week_type_id", None)
                }
            })

        # Dodawanie wydarzeń związanych z godzinami pracy
        for work in work_hours:
            events.append({
                "id": str(work.id),
                "title": "Godziny pracy",
                "start": f"{work.date}T{work.start_time.isoformat()}",
                "end": f"{work.date}T{work.end_time.isoformat()}",
                "color": "#ad0707",
                "textColor": "#FFFFFF",
                "extendedProps": {
                    "type": "work",
                    "date": work.date.isoformat(),
                    "start_time": work.start_time.isoformat(),
                    "end_time": work.end_time.isoformat()
                }
            })

        # Dodaj święta z kalendarza polskiego
        current_year = start_date.year
        polish_holidays = holidays.Poland(years=current_year)

        for date, name in polish_holidays.items():
            if start_date <= date <= end_date:
                events.append({
                    "id": f"holiday-{date}",
                    "title": name,
                    "start": f"{date}T00:00:00",
                    "end": f"{date}T23:59:59",
                    "color": "#9c6802",
                    "textColor": "#FFFFFF",
                    "extendedProps": {
                        "type": "holiday",
                        "date": date.isoformat(),
                        "holiday_name": name
                    }
                })

        return jsonify(events)

    except Exception:
        return jsonify({"error": "Wystąpił błąd podczas pobierania wydarzeń"}), 500


@main.route('/update_work_hours', methods=['POST'])
@login_required
def update_work_hours():
    event_id = request.form.get('event_id')
    date = request.form.get('date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')

    try:
        # Pobranie rekordu godzin pracy i sprawdzenie, czy istnieje i należy do użytkownika
        work_record = WorkHours.query.filter_by(id=event_id, user_id=current_user.id).first()
        if not work_record:
            flash("Rekord nie istnieje lub nie masz do niego dostępu.", 'danger')
            return redirect(url_for('main.index'))

        # Aktualizacja tylko tych pól, które zostały przesłane
        if date:
            work_record.date = datetime.datetime.strptime(date, '%Y-%m-%d').date()

        # Parsowanie start_time i end_time z obsługą różnych formatów
        if start_time:
            work_record.start_time = parse_time(start_time)
        if end_time:
            work_record.end_time = parse_time(end_time)

        # Zatwierdzenie zmian w bazie danych
        db.session.commit()
        flash('Zaktualizowano godziny pracy.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Błąd podczas aktualizacji: {str(e)}", 'danger')

    return redirect(url_for('main.index'))


# Funkcja pomocnicza do obsługi różnych formatów czasu
def parse_time(time_str):
    # Próbujemy najpierw parsować z sekundami
    try:
        return datetime.datetime.strptime(time_str, '%H:%M:%S').time()
    except ValueError:
        # Jeśli brak sekund, spróbujmy bez nich
        return datetime.datetime.strptime(time_str, '%H:%M').time()


@main.route('/update_study_schedule', methods=['POST'])
@login_required
def update_study_schedule():
    event_id = request.form.get('event_id')
    dates = request.form.getlist('dates[]')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    subject_name = request.form.get('subject_name')
    instructor_name = request.form.get('instructor_name') or None
    room = request.form.get('room') or None
    week_type_id = request.form.get('week_type_id') or None
    course_type_id = request.form.get('course_type_id')

    try:
        study_schedule = StudySchedule.query.filter_by(id=event_id, user_id=current_user.id).first()
        if not study_schedule:
            return jsonify({'error': "Rekord nie istnieje lub nie masz do niego dostępu."}), 403

        if dates:
            study_schedule.date = dates[0]
        if start_time:
            study_schedule.start_time = parse_time(start_time)
        if end_time:
            study_schedule.end_time = parse_time(end_time)
        if subject_name:
            study_schedule.subject_name = subject_name
        if instructor_name:
            study_schedule.instructor_name = instructor_name
        if room:
            study_schedule.room = room
        if week_type_id:
            study_schedule.week_type_id = week_type_id
        if course_type_id:
            study_schedule.course_type_id = course_type_id

        db.session.commit()
        return jsonify({'message': 'Zaktualizowano harmonogram zajęć.'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f"Błąd podczas aktualizacji: {str(e)}"}), 500


@main.route('/delete_work_hours', methods=['POST'])
@login_required
def delete_work_hours():
    data = request.get_json()
    event_id = data.get('event_id') if data else None

    if not event_id:
        flash("Nie podano ID rekordu.", 'danger')
        return redirect(url_for('main.index'))

    try:
        # Znajdź rekord godzin pracy, aby upewnić się, że należy do użytkownika
        work_record = WorkHours.query.filter_by(id=event_id, user_id=current_user.id).first()

        if not work_record:
            flash("Rekord nie istnieje lub nie masz do niego dostępu.", 'danger')
            return redirect(url_for('main.index'))

        # Usuń rekord z bazy danych
        db.session.delete(work_record)
        db.session.commit()

        flash("Rekord godzin pracy został usunięty.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Błąd podczas usuwania rekordu: {str(e)}", 'danger')

    return redirect(url_for('main.index'))


@main.route('/delete_study_schedule', methods=['POST'])
@login_required
def delete_study_schedule():
    data = request.get_json()
    event_id = data.get('event_id') if data else None

    if not event_id:
        flash("Nie podano ID rekordu.", 'danger')
        return redirect(url_for('main.index'))

    try:
        # Znajdź rekord harmonogramu zajęć, aby upewnić się, że należy do użytkownika
        study_schedule = StudySchedule.query.filter_by(id=event_id, user_id=current_user.id).first()

        if not study_schedule:
            flash("Rekord nie istnieje lub nie masz do niego dostępu.", 'danger')
            return redirect(url_for('main.index'))

        # Usuń rekord z bazy danych
        db.session.delete(study_schedule)
        db.session.commit()

        flash("Rekord harmonogramu zajęć został usunięty.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Błąd podczas usuwania rekordu: {str(e)}", 'danger')

    return redirect(url_for('main.index'))


@main.route('/export_pdf')
@login_required
def export_pdf():
    month = request.args.get('month', type=int, default=datetime.date.today().month)
    year = request.args.get('year', type=int, default=datetime.date.today().year)
    hourly_rate = request.args.get('hourly_rate', type=float, default=0)

    summary_data = get_summary_data(month, year)
    total_hours = sum(day_data['total_work_hours'] for day_data in summary_data['work_hours_data'])
    total_earnings = total_hours * hourly_rate if hourly_rate else 0

    pdf_buffer = BytesIO()
    p = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, height - 40, f"Podsumowanie godzin pracy - {calendar.month_name[month]} {year}")

    p.setFont("Helvetica", 10)
    table_y = height - 80
    p.drawString(50, table_y, "Dzień")
    p.drawString(150, table_y, "Godziny Pracy")
    p.drawString(300, table_y, "Zarobek (PLN)")

    for day_data in summary_data['work_hours_data']:
        table_y -= 20
        hours_worked = day_data['total_work_hours']
        earnings = hours_worked * hourly_rate if hourly_rate else 0

        if day_data['is_holiday'] or day_data['is_weekend']:
            p.setFillColor(colors.lightgrey)
        else:
            p.setFillColor(colors.black)

        p.drawString(50, table_y, day_data['date'].strftime('%Y-%m-%d'))
        p.drawString(150, table_y, day_data['formatted_work_hours'])
        p.drawString(300, table_y, f"{earnings:.2f} PLN")
        p.setFillColor(colors.black)

        if table_y < 40:
            p.showPage()
            p.setFont("Helvetica", 10)
            table_y = height - 40

    table_y -= 20
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, table_y, "Razem")
    p.drawString(150, table_y, f"{total_hours:.2f}")
    p.drawString(300, table_y, f"{total_earnings:.2f} PLN")

    p.showPage()
    p.save()
    pdf_buffer.seek(0)
    return send_file(pdf_buffer, as_attachment=True, download_name=f"Podsumowanie_{month}_{year}.pdf", mimetype="application/pdf")

@main.route('/export_excel')
@login_required
def export_excel():
    month = request.args.get('month', type=int, default=datetime.date.today().month)
    year = request.args.get('year', type=int, default=datetime.date.today().year)
    hourly_rate = request.args.get('hourly_rate', type=float, default=0)

    summary_data = get_summary_data(month, year)
    total_hours = sum(day_data['total_work_hours'] for day_data in summary_data['work_hours_data'])
    total_earnings = total_hours * hourly_rate if hourly_rate else 0

    data = []
    for day_data in summary_data['work_hours_data']:
        hours_worked = day_data['total_work_hours']
        earnings = hours_worked * hourly_rate if hourly_rate else 0
        data.append({
            'Dzień': day_data['date'].strftime('%Y-%m-%d'),
            'Godziny Pracy': day_data['formatted_work_hours'],
            'Zarobek (PLN)': f"{earnings:.2f}"
        })

    data.append({
        'Dzień': 'Razem',
        'Godziny Pracy': f"{total_hours:.2f}",
        'Zarobek (PLN)': f"{total_earnings:.2f}"
    })

    df = pd.DataFrame(data)
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Podsumowanie Godzin Pracy")

    workbook = writer.book
    worksheet = writer.sheets["Podsumowanie Godzin Pracy"]

    holiday_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    weekend_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_idx, day_data in enumerate(summary_data['work_hours_data'], start=2):
        fill = None
        if day_data['is_holiday']:
            fill = holiday_fill
        elif day_data['is_weekend']:
            fill = weekend_fill

        if fill:
            worksheet[f"A{row_idx}"].fill = fill

    writer.close()
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"Podsumowanie_{month}_{year}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
